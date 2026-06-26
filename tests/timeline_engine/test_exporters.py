"""Tests for export modules."""

from __future__ import annotations

import json
from pathlib import Path

from engine.export.csv_exporter import export_timeline_csv, export_transitions_csv
from engine.export.json_exporter import export_session_json
from engine.export.xlsx_exporter import export_person_summary_xlsx
from engine.types import BehaviourTransition, SessionContext, TimelineBehaviour, TimelineSegment, VideoInfo


def _sample_segments():
    return [
        TimelineSegment(
            person_id=1,
            start_time="00:00:10",
            end_time="00:00:20",
            behaviour=TimelineBehaviour.STANDING,
            confidence=0.87,
            duration_s=10.0,
            start_ms=10_000,
            end_ms=20_000,
        ),
        TimelineSegment(
            person_id=1,
            start_time="00:00:20",
            end_time="00:00:45",
            behaviour=TimelineBehaviour.WALKING,
            confidence=0.92,
            duration_s=25.0,
            start_ms=20_000,
            end_ms=45_000,
        ),
    ]


def test_csv_export(tmp_path: Path):
    path = tmp_path / "person_timeline.csv"
    export_timeline_csv(path, _sample_segments())
    text = path.read_text(encoding="utf-8")
    assert "person_id" in text
    assert "Standing" in text
    assert "Walking" in text
    assert "duration_s" not in text.splitlines()[0]


def test_xlsx_export(tmp_path: Path):
    path = tmp_path / "person_summary.xlsx"
    export_person_summary_xlsx(path, _sample_segments())
    from openpyxl import load_workbook

    wb = load_workbook(path)
    assert "Duration Totals" in wb.sheetnames
    assert "Segments" in wb.sheetnames


def test_json_export(tmp_path: Path):
    path = tmp_path / "session_summary.json"
    context = SessionContext(
        video=VideoInfo("test.mp4", 25.0, 640, 480, 100, 4.0),
        yolo_model="yolo11x.pt",
        pose_model="models/rtmpose-m.onnx",
        device="cpu",
        frame_stride=1,
        person_count=1,
    )
    export_session_json(path, context, _sample_segments(), transitions=[])
    data = json.loads(path.read_text(encoding="utf-8"))
    assert data["session"]["person_count"] == 1
    assert "behaviour_transitions" in data


def test_transitions_csv_export(tmp_path):
    path = tmp_path / "person_transitions.csv"
    export_transitions_csv(
        path,
        [
            BehaviourTransition(
                person_id=1,
                transition_time="00:00:05",
                transition_ms=5000,
                from_behaviour=TimelineBehaviour.WALKING,
                to_behaviour=TimelineBehaviour.STANDING,
                confidence=0.85,
            )
        ],
    )
    text = path.read_text(encoding="utf-8")
    assert "Walking" in text
    assert "Standing" in text
